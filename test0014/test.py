from openpyxl import workbook
import openpyxl
import json



def test():
    book = openpyxl.load_workbook('test.xlsx')
    # book.active = book['jy']
    book.active = 0
    sheet = book.active
    with open('data.txt') as fp:
        data = json.load(fp)

    row = 1
    json.dump()
    for key, value in data.items():
        col = 1
        sheet.cell(row, col, key)
        col += 1
        for v in value:
            sheet.cell(row, col, v)
            col += 1
        row += 1

    book.save('test.xlsx')


if __name__ == '__main__':
    test()
